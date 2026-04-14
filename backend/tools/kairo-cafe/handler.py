import os
import json
import openpyxl


def get_tool_data_path():
    """获取Excel源文件路径"""
    return os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "..", "..", "docs", "开罗咖啡店工具"
    )


def get_cache_dir():
    """获取JSON缓存目录（db/tools/kairo-cafe/，由Docker挂载）"""
    # 相对于项目根目录的 db/tools/kairo-cafe/ 目录
    cache_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "..", "db", "tools", "kairo-cafe"
    )
    cache_dir = os.path.normpath(cache_dir)
    os.makedirs(cache_dir, exist_ok=True)
    return cache_dir


def load_or_create_customer_complaints():
    """加载顾客烦恼数据，首次从Excel读取并缓存为JSON"""
    cache_dir = get_cache_dir()
    cache_path = os.path.join(cache_dir, "customer_complaints.json")

    # 如果缓存存在，直接读取
    if os.path.exists(cache_path):
        with open(cache_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    # 从Excel读取
    data_path = get_tool_data_path()
    xlsx_path = os.path.join(data_path, "顾客烦恼攻略.xlsx")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # 跳过标题行
        if i == 1:
            continue  # 跳过表头行
        if row[0] is None:
            break
        item = {
            "序号": row[0],
            "顾客": row[1],
            "名称": row[2],
            "条件": row[3],
            "报酬": row[4],
            "是否已完成": False
        }
        data.append(item)

    # 保存缓存
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return data


def load_or_create_recipes():
    """加载菜谱大全数据，首次从Excel读取并缓存为JSON"""
    cache_dir = get_cache_dir()
    cache_path = os.path.join(cache_dir, "recipes.json")

    # 如果缓存存在，直接读取
    if os.path.exists(cache_path):
        with open(cache_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    # 从Excel读取
    data_path = get_tool_data_path()
    xlsx_path = os.path.join(data_path, "菜谱大全.xlsx")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # 跳过标题行
        if i == 1:
            continue  # 跳过表头行
        if row[0] is None:
            break
        item = {
            "序号": row[0],
            "饮品": row[1],
            "类别": row[2],
            "食材1": row[3],
            "食材2": row[4],
            "装饰物": row[5],
            "价格": row[6],
            "是否已完成": False
        }
        data.append(item)

    # 保存缓存
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return data


def save_customer_complaints(data):
    """保存顾客烦恼数据到JSON"""
    cache_dir = get_cache_dir()
    cache_path = os.path.join(cache_dir, "customer_complaints.json")
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def save_recipes(data):
    """保存菜谱数据到JSON"""
    cache_dir = get_cache_dir()
    cache_path = os.path.join(cache_dir, "recipes.json")
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


class Handler:
    def get_customer_complaints(self):
        """获取顾客烦恼列表"""
        data = load_or_create_customer_complaints()
        return {"status": "ok", "data": data}

    def get_recipes(self):
        """获取菜谱大全列表"""
        data = load_or_create_recipes()
        return {"status": "ok", "data": data}

    def toggle_complaint_complete(self, index=None):
        """切换顾客烦恼完成状态"""
        if index is None:
            return {"status": "error", "message": "缺少index参数"}

        data = load_or_create_customer_complaints()
        for item in data:
            if item["序号"] == index:
                item["是否已完成"] = not item["是否已完成"]
                save_customer_complaints(data)
                return {"status": "ok", "data": item}
        return {"status": "error", "message": "未找到对应记录"}

    def toggle_recipe_complete(self, index=None):
        """切换菜谱完成状态"""
        if index is None:
            return {"status": "error", "message": "缺少index参数"}

        data = load_or_create_recipes()
        for item in data:
            if item["序号"] == index:
                item["是否已完成"] = not item["是否已完成"]
                save_recipes(data)
                return {"status": "ok", "data": item}
        return {"status": "error", "message": "未找到对应记录"}
